from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
wb = load_workbook(
    '
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

data = Reference(sheet, min_col=min_column+1,
                 max_col=max_column, min_row=min_row, max_row=max_row)
# this section is made to cut out tell the areas that are being measured. By adding a min and max and not perfectly aligned with the sheet,youll need to indicate the starting location on the sheet
category = Reference(sheet, min_col=min_column,
                     max_col=min_column, min_row=min_row+1, max_row=max_row)

barchart = BarChart()  # the parameters of the barchart
barchart.add_data(data, titles_from_data=True)  # the titles
barchart.set_categories(category)  # the catergories
# where the graph will be located on the sheet
sheet.add_chart(barchart, 'B12')
barchart.title = 'Sales by product line'  # The title of the chart
barchart.style = 5  # the style of the graph
wb.save('barchart.xlsx')  # exporting the sheet
